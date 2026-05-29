"""
Excel Agent — writes structured intelligence rows into the openpyxl workbook.
Deduplication is URL-based (md5 of the article URL) for reliability across runs.
"""

import hashlib
import logging
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SECTORS, WORKBOOK_PATH

logger = logging.getLogger(__name__)

COLUMNS = ["Date", "Sector", "Headline", "Summary", "Exporter Implication", "Impact", "Source", "URL"]

COL_WIDTHS = {
    "Date": 12,
    "Sector": 18,
    "Headline": 35,
    "Summary": 50,
    "Exporter Implication": 45,
    "Impact": 10,
    "Source": 22,
    "URL": 45,
}

IMPACT_FILLS = {
    "High": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Medium": PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"),
    "Low": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

_seen_url_hashes: set[str] = set()
_workbook: openpyxl.Workbook | None = None


def _url_hash(url: str) -> str:
    return hashlib.md5(url.strip().encode()).hexdigest()


def _apply_header(ws) -> None:
    ws.row_dimensions[1].height = 20
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def initialize_workbook() -> None:
    """Create the workbook with Master + 10 sector sheets if it doesn't exist."""
    global _workbook
    os.makedirs(os.path.dirname(WORKBOOK_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    # Rename the default sheet to Master
    ws_master = wb.active
    ws_master.title = "Master"
    _apply_header(ws_master)

    for sector in SECTORS:
        ws = wb.create_sheet(title=sector)
        _apply_header(ws)

    wb.save(WORKBOOK_PATH)
    _workbook = wb
    logger.info("Workbook initialized at %s", WORKBOOK_PATH)


def _load_workbook() -> openpyxl.Workbook:
    global _workbook, _seen_url_hashes
    if _workbook is not None:
        return _workbook

    if not os.path.exists(WORKBOOK_PATH):
        initialize_workbook()
        return _workbook

    _workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    # Warm the dedup cache from existing URL column in Master
    ws_master = _workbook["Master"]
    url_col_idx = COLUMNS.index("URL") + 1
    for row in ws_master.iter_rows(min_row=2, values_only=True):
        if row and row[url_col_idx - 1]:
            _seen_url_hashes.add(_url_hash(str(row[url_col_idx - 1])))
    logger.info("Loaded workbook with %d existing rows", ws_master.max_row - 1)
    return _workbook


def _append_row(ws, row_dict: dict) -> None:
    impact = row_dict.get("impact", "Low")
    fill = IMPACT_FILLS.get(impact)
    values = [
        row_dict.get("date", ""),
        row_dict.get("sector", ""),
        row_dict.get("headline", ""),
        row_dict.get("summary", ""),
        row_dict.get("exporter_implication", ""),
        impact,
        row_dict.get("source", ""),
        row_dict.get("url", ""),
    ]
    ws.append(values)
    if fill:
        last_row = ws.max_row
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=last_row, column=col_idx).fill = fill


def write_row(row_dict: dict) -> bool:
    """
    Append row_dict to Master and the matching sector sheet.
    Returns True if written, False if duplicate or sector sheet missing.
    """
    wb = _load_workbook()
    url = row_dict.get("url", "")
    h = _url_hash(url) if url else ""

    if h and h in _seen_url_hashes:
        logger.debug("Duplicate skipped: %s", row_dict.get("headline", url))
        return False

    sector = row_dict.get("sector", "")
    if sector not in wb.sheetnames:
        logger.warning("Unknown sector sheet '%s' — writing to Master only", sector)

    _append_row(wb["Master"], row_dict)
    if sector in wb.sheetnames:
        _append_row(wb[sector], row_dict)

    if h:
        _seen_url_hashes.add(h)

    wb.save(WORKBOOK_PATH)
    logger.info("Wrote row: [%s] %s", sector, row_dict.get("headline", "")[:60])
    return True


def get_row_count() -> int:
    """Return the number of data rows in the Master sheet (excluding header)."""
    wb = _load_workbook()
    return max(0, wb["Master"].max_row - 1)
